from openpyxl.workbook import workbook
import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook


def make_tables(files, output_path):
	'''
	Loops through files, reads file.file_name in PMSLT_tables/drawAgg, applies
	transformations to dataframe and saves to excel spreadsheet specified
	by output_path.
	'''
	init_workbook(output_path)
	path_prefix= "PMSLT_tables/drawAgg"
	for file in files:
		df = pd.read_csv(f"{path_prefix}/{file['file_name']}")

		
		# Make population column
		df["Population"] = df[["Sex", "Strata"]].agg(" ".join, axis=1)
		df.loc[df["Population"] == "All All", "Population"] = "All"

		# Drop columns and rename
		df.drop(["BAU","fullcease","fullceasevape", "Sex", "Strata"], axis=1 ,inplace=True)
		df.rename({
			"ENDS in pharmacies only": "Intervention 1",
			"ENDS in pharmacies only + advice to quit smoking": "Intervention 2",
			"sensitivity on effect of advice": "Intervention 2.1",
			"ENDS in pharmacies only + advice to quit vaping": "Intervention 3",
			"ENDS in pharmacies + GP prescription" : "Intervention 4",
		},
		axis="columns", inplace=True)
		df.set_index(["Population", "Year"], inplace=True)

		# Split each column into "Estimate" and "Uncertainty"
		level_0_cols = df.columns
		df.columns = pd.MultiIndex.from_product([df.columns, ["Estimate"]])
		for column in level_0_cols:
			df[column, "95% UI"] = df[column, "Estimate"].str.split("(").map(lambda x: f"({x[1]}")
			df[column, "Estimate"] = df[column, "Estimate"].str.split("(").map(lambda x: float(x[0].strip().replace(",", "")))

		# Column ordering
		level_0_order = [
			"Intervention 1","Intervention 2","Intervention 2.1",
			"Intervention 3","Intervention 4"
		]
		level_1_order = ["Estimate", "95% UI"]
		df = df.reindex(columns=level_0_order, level=0)
		df = df.reindex(columns=level_1_order, level=1)

		# Save file
		save_file(
			df=df,
			sheet_name=file["title"],
			output_path=output_path
		)

	clean_workbook(output_path)


def init_workbook(output_path):
	'''
	Initialises empty workbook	
	'''
	book = Workbook()
	book.save(filename = output_path)
	
def clean_workbook(output_path):
	'''
	Removes empty sheet to avoid file corruption.
	'''
	book = load_workbook(output_path)
	book.remove_sheet(book["Sheet"])
	book.save(filename = output_path)

			
def save_file(df, sheet_name, output_path):
	'''
	Writes dataframe to specified sheet in output excel file.
	'''
	book = load_workbook(output_path)
	writer = pd.ExcelWriter(output_path, engine = 'openpyxl')
	writer.book = book
	writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
	# writer = pd.ExcelWriter("PMSLT_tables/output/output.xlsx", engine="openpyxl")
	print(sheet_name)
	df.to_excel(writer, sheet_name=sheet_name)
	writer.save()
	writer.close()


files = [
	{"file_name": "out_deaths_year_year_0-111_discount_0.csv", "title": "deaths_discount_0"}, 
	{"file_name": "out_HALY_year_year_0-111_discount_-0.03.csv", "title": "HALYS_discount_0.03"}, 
	{"file_name": "out_HALY_year_year_0-111_discount_0.csv", "title": "HALYS_discount_0"}, 
	{"file_name": "out_total_income_year_year_0-111_discount_-0.03_millions.csv", "title": "income_discount_0.03"}, 
	{"file_name": "out_total_income_year_year_0-111_discount_0_millions.csv", "title": "income_discount_0"}, 
	{"file_name": "out_total_spent_year_year_0-111_discount_-0.03_millions.csv", "title": "expenditure_discount_0.03"}, 
	{"file_name": "out_total_spent_year_year_0-111_discount_0_millions.csv", "title": "expenditure_discount_0"}, 
]


make_tables(
	files=files,
	output_path="PMSLT_tables/output/vaping_tables.xlsx"
)
print("done")
